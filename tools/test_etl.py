import csv as _csv_mod
import os
import tempfile
import unittest
from mel_etl.normalize import limpiar_celda, a_entero, solo_fecha, normalizar_sexo, normalizar_enum
from mel_etl.extract import leer_hoja, escribir_csv, extraer_dimensiones, reasignar_ids, _TRANSFORMS
from mel_etl.extract import extraer_cadena_programada, extraer_ejecuciones_y_participacion
import openpyxl

XLSX = os.path.join(os.path.dirname(__file__), "..", "CPJ_MEL_v1_9_seguimiento_actualizado (12).xlsx")


class TestNormalize(unittest.TestCase):
    def test_limpiar_celda_errores_y_vacios(self):
        self.assertEqual(limpiar_celda(None), "")
        self.assertEqual(limpiar_celda("  "), "")
        self.assertEqual(limpiar_celda("#REF!"), "")
        self.assertEqual(limpiar_celda("#¡REF!"), "")
        self.assertEqual(limpiar_celda("#N/A"), "")
        self.assertEqual(limpiar_celda("  Hola  "), "Hola")

    def test_a_entero(self):
        self.assertEqual(a_entero("1966.0"), "1966")
        self.assertEqual(a_entero("6561566319.0"), "6561566319")
        self.assertEqual(a_entero(1966), "1966")
        self.assertEqual(a_entero(""), "")
        self.assertEqual(a_entero("abc"), "")

    def test_solo_fecha(self):
        self.assertEqual(solo_fecha("2026-05-02 00:00:00"), "2026-05-02")
        self.assertEqual(solo_fecha("2026-05-02"), "2026-05-02")
        self.assertEqual(solo_fecha(""), "")

    def test_normalizar_sexo(self):
        self.assertEqual(normalizar_sexo("M"), "F")   # Excel M = Mujer
        self.assertEqual(normalizar_sexo("H"), "M")   # Excel H = Hombre
        self.assertEqual(normalizar_sexo("N"), "X")
        self.assertEqual(normalizar_sexo("m"), "F")
        self.assertEqual(normalizar_sexo(""), "")

    def test_normalizar_enum(self):
        validos = ["programado", "ejecutado", "cancelado", "reprogramado"]
        self.assertEqual(normalizar_enum("EJECUTADO", {}, validos), "ejecutado")
        self.assertEqual(normalizar_enum("REPROGRAMADO", {}, validos), "reprogramado")
        self.assertEqual(normalizar_enum("desconocido", {}, validos), "")
        self.assertEqual(normalizar_enum("ACTIVO", {"activo": "activo"}, ["activo", "inactivo"]), "activo")


class TestDimensiones(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    def test_leer_hoja_dim_ejes(self):
        filas = leer_hoja(self.wb, "🔵dim_ejes")
        self.assertGreaterEqual(len(filas), 3)           # ≥3 ejes reales
        self.assertIn("id_eje", filas[0])
        self.assertTrue(filas[0]["id_eje"].startswith("EJE_"))

    def test_extraer_dimensiones_genera_5_csv(self):
        with tempfile.TemporaryDirectory() as d:
            conteos = extraer_dimensiones(self.wb, d)
            for nombre in ["ejes", "instituciones", "lineas", "componentes", "actividades"]:
                self.assertTrue(os.path.exists(os.path.join(d, f"{nombre}.csv")), f"falta {nombre}.csv")
                self.assertGreater(conteos[nombre], 0, f"{nombre} vacío")
            self.assertEqual(conteos["instituciones"], 5)
            # cabecera de actividades.csv = columnas del esquema
            with open(os.path.join(d, "actividades.csv")) as fh:
                cab = fh.readline().strip().split(",")
            self.assertEqual(cab[:3], ["id_actividad", "num_actividad", "nombre"])


class TestReasignarIds(unittest.TestCase):
    def test_reasignar_ids_secuencial(self):
        filas = [{"id": "PROC_00001"}, {"id": "PROC_00005"}, {"id": ""}, {"id": "PROC_00009"}]
        mapa = reasignar_ids(filas, "id")
        self.assertEqual(mapa, {"PROC_00001": 1, "PROC_00005": 2, "PROC_00009": 3})

    def test_reasignar_ids_sin_duplicar(self):
        filas = [{"id": "EVP_1"}, {"id": "EVP_1"}, {"id": "EVP_2"}]
        mapa = reasignar_ids(filas, "id")
        self.assertEqual(mapa, {"EVP_1": 1, "EVP_2": 2})


class TestCompIdTransform(unittest.TestCase):
    """Tests para la transform comp_id: COM_NNN → COMP_NNN."""

    def setUp(self):
        self.fn = _TRANSFORMS["comp_id"]

    def test_com_normaliza_a_comp(self):
        self.assertEqual(self.fn("COM_006"), "COMP_006")

    def test_comp_permanece_igual(self):
        self.assertEqual(self.fn("COMP_006"), "COMP_006")

    def test_vacio_permanece_vacio(self):
        self.assertEqual(self.fn(""), "")
        self.assertEqual(self.fn(None), "")

    def test_com_con_ceros_multiples(self):
        self.assertEqual(self.fn("COM_012"), "COMP_012")

    def test_no_normaliza_prefijos_distintos(self):
        # COMP_ ya es correcto, COM_ sin dígitos no debe transformarse a medias
        self.assertEqual(self.fn("COMP_999"), "COMP_999")
        # Cadena sin patrón COM_NNN queda intacta
        self.assertEqual(self.fn("ACT_001"), "ACT_001")


class TestSinFKDangling(unittest.TestCase):
    """Test de integración: tras extraer_dimensiones + extraer_cadena_programada,
    ningún id_actividad en eventos.csv puede quedar fuera del set de actividades.csv.
    """

    @classmethod
    def setUpClass(cls):
        cls.wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    def test_no_fk_dangling_eventos(self):
        with tempfile.TemporaryDirectory() as d:
            conteos = extraer_dimensiones(self.wb, d)
            act_ids = conteos["act_ids"]
            self.assertGreater(len(act_ids), 0, "act_ids vacío")

            extraer_cadena_programada(self.wb, d, act_ids=act_ids)

            # Verificar que todos los id_actividad de eventos.csv están en act_ids.
            with open(os.path.join(d, "eventos.csv"), newline="", encoding="utf-8") as fh:
                filas_ev = list(_csv_mod.DictReader(fh))
            dangling = [
                r["id_actividad"]
                for r in filas_ev
                if r.get("id_actividad") and r["id_actividad"] not in act_ids
            ]
            self.assertEqual(dangling, [], f"FK dangling en eventos.csv: {dangling[:5]}")

    def test_no_fk_dangling_procesos(self):
        with tempfile.TemporaryDirectory() as d:
            conteos = extraer_dimensiones(self.wb, d)
            act_ids = conteos["act_ids"]

            extraer_cadena_programada(self.wb, d, act_ids=act_ids)

            with open(os.path.join(d, "procesos.csv"), newline="", encoding="utf-8") as fh:
                filas_proc = list(_csv_mod.DictReader(fh))
            dangling = [
                r["id_actividad"]
                for r in filas_proc
                if r.get("id_actividad") and r["id_actividad"] not in act_ids
            ]
            self.assertEqual(dangling, [], f"FK dangling en procesos.csv: {dangling[:5]}")


class TestCadenaMEL(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    def test_participaciones_y_sexo(self):
        with tempfile.TemporaryDirectory() as d:
            cad = extraer_cadena_programada(self.wb, d)
            res = extraer_ejecuciones_y_participacion(self.wb, d, cad["mapa_eventos"])
            for nombre in ["ejecuciones", "participaciones", "agregadas"]:
                self.assertTrue(os.path.exists(os.path.join(d, f"{nombre}.csv")))
            self.assertGreater(res["participaciones"], 800)   # línea base ≈988
            # sexo normalizado: solo F/M/X o vacío
            import csv as _csv
            with open(os.path.join(d, "participaciones.csv")) as fh:
                sexos = {row["sexo"] for row in _csv.DictReader(fh)}
            self.assertTrue(sexos <= {"F", "M", "X", ""}, f"sexo sin normalizar: {sexos}")
            # id_ejecucion es entero
            with open(os.path.join(d, "participaciones.csv")) as fh:
                primera = next(_csv.DictReader(fh))
            self.assertTrue(primera["id_ejecucion"].isdigit() or primera["id_ejecucion"] == "")


if __name__ == "__main__":
    unittest.main()
